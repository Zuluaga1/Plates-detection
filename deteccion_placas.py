# -*- coding: utf-8 -*-
"""
Created on Wed May 12 10:39:57 2021

@author: esteb
PROYECTO FINAL
"""
import cv2
import pytesseract
import pandas as pd
#import openpyxl
from openpyxl import load_workbook
import pandas.io.formats.excel
#from datetime import datetime
#from datetime import datetime, timedelta
#import xlsxwriter
import datetime
from src.database import *
import qrcode

pytesseract.pytesseract.tesseract_cmd = r'D:/tesseract/tesseract'
base = DataBase()
cap = cv2.VideoCapture(0,cv2.CAP_DSHOW) # mostrar toda la ventana
placa = []
n = True
m = [0,0]
s = 0 #0 si entra 1 si va saliendo
#def main():

while n == True:
    #global probando
    #global probando1
    ret,frame = cap.read()
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    gray = cv2.blur(gray,(3,3)) #atenuando el ruido, como quitando algunos bordes
    canny = cv2.Canny(gray,150,200) #deteccion de bordes
    canny = cv2.dilate(canny,None,iterations=1) #engrozar los bordes
    #_,cnts,_ = cv2.findContours(canny,cv2.RETR_LIST,cv2.CHAIN_APPROX_SIMPLE)
    cnts,_ = cv2.findContours(canny,cv2.RETR_LIST,cv2.CHAIN_APPROX_SIMPLE)
    
    for c in cnts:
      area = cv2.contourArea(c) #detectar área
      x,y,w,h = cv2.boundingRect(c)    #detectando rectángulo
      epsilon = 0.09*cv2.arcLength(c,True) #9 % es con experimentación
      approx = cv2.approxPolyDP(c,epsilon,True) #approx es para ver que tiene 4 vértices
      
      if len(approx)== 4 and area>9000:   
          cv2.drawContours(frame,[c],0,(0,255,0),2)
          #cv2.drawContours(image,[approx],0,(0,255,0),3)
          aspect_ratio = float(w)/h
          #global probando
          if aspect_ratio>1.5:
              global probando
              global probando1
              placa = gray[y:y+h,x:x+w] #área donde está presente la placa gris
              text = pytesseract.image_to_string(placa)
              probando = text.split()
              quitar = "{}()-,:;.\n!\"'\ "
              for caracter in quitar: #quitar espacios y caracteres extraños
                  text = text.replace(caracter, "")
              probando1 = text.split()
              texto = ' '.join([str(item) for item in probando1])
              print('PLACA: ',text)
              cv2.imshow('PLACA',placa)
              cv2.moveWindow('PLACA',780,10) #pos de la placa
              cv2.rectangle(frame,(x,y),(x+w,y+h),(0,255,0),3)
              cv2.putText(frame,text,(x-20,y-10),1,2.2,(0,255,0),3)
              if (s == 0): #entrando
                #if (probando[0].isalpha() == True) & (probando[1].isnumeric()) & (len(probando[1])==3) & len(probando[0])==3:
                #print("alooooooooooooo")
                numero_placa = base.select_all() #read db
                numero_placa = pd.DataFrame(numero_placa,columns=("index",
                                                                  "PLACA", "FECHA"))
                #entrada = pd.read_excel('carros.xlsx')
                blanca = pd.read_excel('blanca.xlsx')
                placa_blanca = blanca['PLACA'] == texto
                #placa_blanca = numero_placa['PLACA'] == texto
                if (placa_blanca == True).any():
                    print('lista blanca')
                    n = False
                    #pos = entrada.iloc[0][0]
                    #pos1 = pos + 1
                    time = datetime.datetime.now()
                    valor = numero_placa['PLACA'] == texto
                    if (valor == True).any():
                        print("no se agrega a la base de datos")
                    else:
                        print("si agrega a la base de datos")
                        base.insert(texto,time) #insert into db
                        formato = texto[:3]+" "+texto[-3:]
                        data = "El usuario con placa {} ha guardado el auto en la fecha {}".format(str(formato),time)
                        imagen = qrcode.make(data)
                        imagen.save("qrcode.jpg")
                        
                        """
                        df = pd.DataFrame({"placa":probando1, 
                                           "time":time})
                        
                        book = load_workbook('carros.xlsx')
                        writer = pd.ExcelWriter('carros.xlsx', engine='openpyxl') 
                        writer.book = book
                        ws = book.active #para escribir en el excel
                        pandas.io.formats.excel.header_style = None
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, book.worksheets[0].title,startcol=1,
                                    index = False, startrow=pos1, header=False)
                        ws['A2'] = pos1
                        writer.save()
                        """
                            
              elif s == 1:  #saliendo
               
                blanca = pd.read_excel('blanca.xlsx')
                #salida = pd.DataFrame(salida)
                time1 = datetime.datetime.now()
                numero_placa = base.select_all() #read db
                numero_placa = pd.DataFrame(numero_placa,columns=("index",
                                                                  "PLACA", "FECHA"))
                numero_placa_sal = base.select_all_sal()
                numero_placa_sal = pd.DataFrame(numero_placa_sal,columns=("index",
                                                                  "PLACA", "FECHA"))
                comparador = numero_placa_sal['PLACA'] == texto
                valor1 = blanca['PLACA'] == texto
                if (valor1 == True).any():
                    n = False       
                    print('Vehículo Saliendo')
                    if (comparador == True).any():
                        print("no se agrega a la base de datos de salida")
                    else:
                        print('si se agrega salida')
                        base.insert_sal(texto,time1)
                        
                        resta = numero_placa[numero_placa['PLACA'] ==texto]
                        resta = time1 - resta['FECHA']
                        
                        #hula ="xdddd"
                        #print(hula)
                        # resta = salida.loc[salida['PLACA']==texto]
                        # tiempo = time1 - resta['DATETIME']
                        tiempo = resta.to_string()
                        tiempo = tiempo.split(" ")
                        #a1 =datetime.now()
                        #a = datetime.strptime(tiempo[-1],'%H:%M:%S.%f').timestamp()
                        a = datetime.datetime.strptime(tiempo[-1],"%H:%M:%S.%f")
                        date = a - datetime.datetime(1900, 1, 1)
                        seconds = date.total_seconds()
                        factura =(seconds * 2000)/3600
                        total = round(factura,0)
                        
                        print("el tiempo del vehículo con placas {} fue de {}".format(texto,tiempo[-1])) 
                        print("El usuario deberá pagar un total de ",total)
                      
    cv2.imshow('frame', frame)
    k = cv2.waitKey(10)
        #if k == 5:
         #   break
cap.release()
cv2.destroyAllWindows()
